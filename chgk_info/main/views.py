from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from rest_framework import viewsets
from rest_framework.parsers import JSONParser
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.utils import json
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse, JsonResponse, Http404
from .models import *
from .serializers import *
from openpyxl import load_workbook
# Create your views here.


@csrf_exempt
def cities_list(request):
    if request.method == 'GET':
        cities = City.objects.all()
        serializer = CitySerializer(cities, many=True)
        return JsonResponse(serializer.data, safe=False)
    elif request.method == 'POST':
        data = JSONParser().parse(request)
        serializer = CitySerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)
        return JsonResponse(serializer.errors, status=400)


class CityDetailView(APIView):
    def get_object(self, pk):
        try:
            return City.objects.get(pk=pk)
        except City.DoesNotExist:
            raise Http404

    def get(self, request, pk):
        city = self.get_object(pk)
        serializer = CitySerializer(city)
        return Response(serializer.data)

    def delete(self, request, pk):
        city = self.get_object(pk)
        city.delete()
        return Response(status=204)


@csrf_exempt
def teams_list(request):
    if request.method == 'GET':
        teams = Team.objects.all()
        serializer = TeamSerializer(teams, many=True)
        return JsonResponse(serializer.data, safe=False)
    elif request.method == 'POST':
        data = JSONParser().parse(request)
        serializer = TeamSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)
        return JsonResponse(serializer.errors, status=400)


class TeamDetailView(APIView):
    def get_object(self, pk):
        try:
            return Team.objects.get(pk=pk)
        except Team.DoesNotExist:
            raise Http404

    def get(self, request, pk):
        team = self.get_object(pk)
        serializer = TeamSerializer(team)
        return Response(serializer.data)

    def delete(self, request, pk):
        team = self.get_object(pk)
        team.delete()
        return Response(status=204)


class PlayerListView(APIView):
    def get(self, request):
        players = Player.objects.all()
        serializer = PlayerSerializer(players, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = PlayerSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)


class PlayerDetailView(APIView):
    def get_object(self, pk):
        try:
            return Player.objects.get(pk=pk)
        except Player.DoesNotExist:
            raise Http404

    def get(self, request, pk):
        player = self.get_object(pk)
        serializer = PlayerSerializer(player)
        return Response(serializer.data)

    def delete(self, request, pk):
        player = self.get_object(pk)
        player.delete()
        return Response(status=204)


class ChgkUserViewSet(viewsets.ViewSet):
    def create(self, request):
        serializer = ChgkUserRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def list(self, request):
        queryset = ChgkUser.objects.all()
        serializer = ChgkUserSerializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, pk):
        queryset = ChgkUser.objects.all()
        user = get_object_or_404(queryset, pk=pk)
        serializer = ChgkUserSerializer(user)
        return Response(serializer.data)


class TournamentBaseViewSet(viewsets.ViewSet):
    def list(self, request):
        queryset = TournamentBaseModel.objects.all()
        serializer = TournamentBaseModelSerializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, pk):
        queryset = TournamentBaseModel.objects.all()
        tournament = get_object_or_404(queryset, pk=pk)
        serializer = TournamentBaseModelSerializer(tournament)
        return Response(serializer.data)

    def delete(self, request, pk):
        queryset = TournamentBaseModel.objects.all()
        tournament = get_object_or_404(queryset, pk=pk)
        tournament.deleted = True
        tournament.save()
        return Response(status=204)
        # tournament.delete()
        # return Response(status=204)


class SynchronousViewSet(viewsets.ViewSet):
    def list(self, request):
        queryset = Synchronous.objects.all()
        serializer = SynchronousSerializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, pk):
        queryset = Synchronous.objects.all()
        synchron = get_object_or_404(queryset, pk=pk)
        serializer = TournamentBaseModelSerializer(synchron)
        return Response(serializer.data)

    def create(self, request):
        serializer = SynchronousSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, pk):
        queryset = Synchronous.objects.all()
        synchron = get_object_or_404(queryset, pk=pk)
        synchron.delete()
        return Response(status=204)


class CupViewSet(viewsets.ViewSet):
    permission_classes_by_action = {'create': [IsAuthenticated], 'list': [AllowAny],
                                    'retrieve': [AllowAny], 'delete': [IsAuthenticated]}

    def list(self, request):
        queryset = Cup.objects.all()
        serializer = CupSerializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, pk):
        queryset = Cup.objects.all()
        cup = get_object_or_404(queryset, pk=pk)
        serializer = TournamentBaseModelSerializer(cup)
        return Response(serializer.data)

    def create(self, request):
        user = request.user
        request.data['posted_by'] = user.id
        serializer = CupSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, pk):
        queryset = Cup.objects.all()
        cup = get_object_or_404(queryset, pk=pk)
        cup.delete()
        return Response(status=204)


class ApplicationViewSet(viewsets.ViewSet):
    permission_classes_by_action = {'create': [IsAuthenticated], 'list': [AllowAny]}

    def create(self, request, pk):
        queryset = Synchronous.objects.all()
        synchron = get_object_or_404(queryset, pk=pk)
        user = request.user
        request.data["representative"] = user.id
        request.data["synchron"] = pk
        serializer = ApplicationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def list(self, request, pk):
        queryset = Application.objects.filter(synchron=pk)
        serializer = ApplicationSerializer(queryset, many=True)
        return Response(serializer.data)


class ProfileViewSet(viewsets.ViewSet):
    permission_classes = (IsAuthenticated,)

    def retrieve(self, request):
        app_queryset = Application.objects.select_related().filter(host=request.user.id)
        app_serializer = ApplicationSerializer(app_queryset, many=True)
        synchron_ids = []
        for x in app_serializer.data:
            synchron_ids.append(x['synchron'])
        sync_queryset = Synchronous.objects.filter(id__in = synchron_ids)
        sync_serializer = SynchronousSerializer(sync_queryset, many=True)
        user_queryset = ChgkUser.objects.get(id=request.user.id)
        user_serializer = ChgkUserSerializer(user_queryset)
        y = user_serializer.data
        y['files'] = [x['question_file'] for x in sync_serializer.data]
        return Response(y)


def file_parser(request, tournament):
    file = load_workbook(request.FILES['results'])
    score_sheet = file.worksheets[0]
    row = 1
    score_by_team = {}
    while True:
        if score_sheet[f'A{row}'].value is None:
            break
        else:
            team_id = score_sheet[f'A{row}'].value
            team_name = score_sheet[f'B{row}'].value
            team_city = score_sheet[f'C{row}'].value
            tour_no = score_sheet[f'D{row}'].value
            if score_by_team.get(team_name) is None:
                score_by_team[team_name] = {
                    'id': team_id,
                    'tct_id': None,
                    'city': team_city,
                    'results': {
                        'total': 0
                    }
                }
            column = 4
            while True:
                value = score_sheet[f'''{chr(column + ord('A'))}{row}'''].value
                if value is None:
                    break
                else:
                    if score_by_team[team_name]['results'].get(tour_no) is None:
                        score_by_team[team_name]['results'][tour_no] = 0
                    score_by_team[team_name]['results'][tour_no] += value
                    score_by_team[team_name]['results']['total'] += value
                column += 1
        row += 1
    for name, val in score_by_team.items():
        if val.get('id') == 0:
            try:
                city = City.objects.get(name=val.get('city'))
            except City.DoesNotExist:
                city = City.objects.create(name=val.get('city'))
            new_team = Team.objects.create(name=name, rating=0, city=city)
            val['id'] = new_team.id
        team = Team.objects.get(id=val['id'])
        tct = TournamentCompetitorsTeams.objects.create(
            tournament=tournament,
            team=team,
            alias_name=(name if team.name != name else None),
            results=val['results']
        )
        score_by_team[name]['tct_id'] = tct.id

    file = load_workbook(request.FILES['teams'])
    teams_sheet = file.worksheets[0]
    row = 1
    while True:
        if teams_sheet[f'A{row}'].value is None:
            break
        else:
            team_id = teams_sheet[f'A{row}'].value
            team_name = teams_sheet[f'B{row}'].value
            team_city = teams_sheet[f'C{row}'].value
            player_type = teams_sheet[f'D{row}'].value
            player_id = teams_sheet[f'E{row}'].value
            player_first_name = teams_sheet[f'F{row}'].value
            player_last_name = teams_sheet[f'G{row}'].value
            if team_id == 0:
                team_id = score_by_team[team_name]['id']
            try:
                city = City.objects.get(name=team_city)
            except City.DoesNotExist:
                city = City.objects.create(name=team_city)
            player = None
            if player_id == 0:
                team = Team.objects.get(id=team_id)
                if player_type == 'Л':
                    city = None
                    team = None
                player = Player.objects.create(
                    firstname=player_first_name,
                    lastname=player_last_name,
                    city=city,
                    team=team,
                    rating=0
                )
            else:
                player = Player.objects.get(id=player_id)
            tct = TournamentCompetitorsTeams.objects.get(id=score_by_team[team_name]['tct_id'])
            tcp = TournamentCompetitorsPlayers.objects.create(
                tournament_team=tct,
                player=player
            )
        row += 1
    return Response(score_by_team)


class SynchronUploadViewSet(viewsets.ViewSet):
    permission_classes = (IsAuthenticated,)

    def create(self, request, pk):
        queryset = Synchronous.objects.all()
        synchron = get_object_or_404(queryset, pk=pk)
        representative = request.user
        queryset = Application.objects.all()
        application = get_object_or_404(queryset, synchron=synchron, representative=representative, status=False)
        Application.objects.filter(synchron=synchron, representative=representative, status=False).update(status=True)
        return file_parser(request, synchron)


class CupUploadViewSet(viewsets.ViewSet):
    permission_classes = (IsAuthenticated,)

    def create(self, request, pk):
        queryset = Cup.objects.filter(posted_by=request.user)
        cup = get_object_or_404(queryset, pk=pk)
        return file_parser(request, cup)