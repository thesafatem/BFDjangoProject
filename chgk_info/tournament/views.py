from django.http import Http404
from django.shortcuts import get_object_or_404
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from openpyxl import load_workbook
from .models import TournamentBaseModel, Regular, Cup, Application, \
    TournamentCompetitorsTeams, TournamentCompetitorsPlayers
from .serializers import TournamentBaseModelSerializer, RegularSerializer, CupSerializer, \
    ApplicationSerializer, TournamentCompetitorsTeamsSerializer
from team.models import Team
from city.models import City
from player.models import Player

import logging
logger = logging.getLogger('tournament')
# Create your views here.


class TournamentBaseViewSet(viewsets.ViewSet):
    permission_classes_by_action = {'list': [AllowAny], 'retrieve': [AllowAny], 'delete': [IsAuthenticated]}

    def get_serialized_object(self, pk):
        try:
            return RegularSerializer(Regular.objects.get(pk=pk))
        except Regular.DoesNotExist:
            try:
                return CupSerializer(Cup.objects.get(pk=pk))
            except Cup.DoesNotExist:
                raise Http404

    def list(self, request):
        queryset = TournamentBaseModel.objects.all()
        serializer = TournamentBaseModelSerializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, pk):
        serializer = self.get_serialized_object(pk)
        return Response(serializer.data)

    def delete(self, request, pk):
        queryset = TournamentBaseModel.objects.all()
        tournament = get_object_or_404(queryset, pk=pk)
        user = request.user
        if tournament.posted_by != user:
            logger.error(f'User (id = {user.id}) has no right to delete tournament (id = {tournament.id}, name = {tournament.name})')
            return Response(status=403)
        tournament.deleted = True
        tournament.save()
        logger.info(f'Tournament (id = {tournament.id}) deleted')
        return Response(status=204)


class RegularViewSet(viewsets.ViewSet):
    permission_classes_by_action = {'create': [IsAuthenticated], 'list': [AllowAny]}

    def list(self, request):
        queryset = Regular.objects.all()
        serializer = RegularSerializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request):
        user = request.user
        request.data['posted_by'] = user.id
        serializer = RegularSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info(f'Regular tournament (id = {serializer.data["id"]}) created')
            return Response(serializer.data)
        logger.error(serializer.errors)
        return Response(serializer.errors, status=400)


class CupViewSet(viewsets.ViewSet):
    permission_classes_by_action = {'create': [IsAuthenticated], 'list': [AllowAny]}

    def list(self, request):
        queryset = Cup.objects.all()
        serializer = CupSerializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request):
        user = request.user
        request.data['posted_by'] = user.id
        serializer = CupSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info(f'Cup tournament (id = {serializer.data["id"]}) created')
            return Response(serializer.data)
        logger.error(serializer.errors)
        return Response(serializer.errors, status=400)


class ApplicationViewSet(viewsets.ViewSet):
    permission_classes_by_action = {'create': [IsAuthenticated], 'list': [AllowAny]}

    def create(self, request, pk):
        queryset = Regular.objects.all()
        regular = get_object_or_404(queryset, pk=pk)
        user = request.user
        request.data["representative"] = user.id
        request.data["regular"] = pk
        serializer = ApplicationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info(f'User (id = {user.id}) commit application for the tournament (id = {regular.id})')
            return Response(serializer.data)
        logger.error(serializer.errors)
        return Response(serializer.errors, status=400)

    def list(self, request, pk):
        queryset = Application.objects.filter(regular=pk)
        serializer = ApplicationSerializer(queryset, many=True)
        return Response(serializer.data)


def file_parser(request, tournament):
    file = load_workbook(request.FILES['results'])
    score_sheet = file.worksheets[0]
    row = 1
    score_by_team = {}
    while True:
        if score_sheet[f'A{row}'].value is None:
            break
        else:
            team_id = score_sheet[f'A{row}'].value
            team_name = score_sheet[f'B{row}'].value
            team_city = score_sheet[f'C{row}'].value
            tour_no = score_sheet[f'D{row}'].value
            if score_by_team.get(team_name) is None:
                score_by_team[team_name] = {
                    'id': team_id,
                    'tct_id': None,
                    'city': team_city,
                    'results': {
                        'total': 0
                    }
                }
            column = 4
            while True:
                value = score_sheet[f'''{chr(column + ord('A'))}{row}'''].value
                if value is None:
                    break
                else:
                    if score_by_team[team_name]['results'].get(tour_no) is None:
                        score_by_team[team_name]['results'][tour_no] = 0
                    score_by_team[team_name]['results'][tour_no] += value
                    score_by_team[team_name]['results']['total'] += value
                column += 1
        row += 1
    for name, val in score_by_team.items():
        if val.get('id') == 0:
            try:
                city = City.objects.get(name=val.get('city'))
            except City.DoesNotExist:
                city = City.objects.create(name=val.get('city'))
            new_team = Team.objects.create(name=name, rating=0, city=city)
            val['id'] = new_team.id
        team = Team.objects.get(id=val['id'])
        tct = TournamentCompetitorsTeams.objects.create(
            tournament=tournament,
            team=team,
            alias_name=(name if team.name != name else None),
            results=val['results']
        )
        score_by_team[name]['tct_id'] = tct.id

    file = load_workbook(request.FILES['teams'])
    teams_sheet = file.worksheets[0]
    row = 1
    while True:
        if teams_sheet[f'A{row}'].value is None:
            break
        else:
            team_id = teams_sheet[f'A{row}'].value
            team_name = teams_sheet[f'B{row}'].value
            team_city = teams_sheet[f'C{row}'].value
            player_type = teams_sheet[f'D{row}'].value
            player_id = teams_sheet[f'E{row}'].value
            player_first_name = teams_sheet[f'F{row}'].value
            player_last_name = teams_sheet[f'G{row}'].value
            if team_id == 0:
                team_id = score_by_team[team_name]['id']
            try:
                city = City.objects.get(name=team_city)
            except City.DoesNotExist:
                city = City.objects.create(name=team_city)
            player = None
            if player_id == 0:
                team = Team.objects.get(id=team_id)
                if player_type == 'Л':
                    city = None
                    team = None
                player = Player.objects.create(
                    first_name=player_first_name,
                    last_name=player_last_name,
                    city=city,
                    team=team,
                    rating=0
                )
            else:
                player = Player.objects.get(id=player_id)
            tct = TournamentCompetitorsTeams.objects.get(id=score_by_team[team_name]['tct_id'])
            tcp = TournamentCompetitorsPlayers.objects.create(
                tournament_team=tct,
                player=player
            )
        row += 1
    return Response(score_by_team)


class RegularUploadViewSet(viewsets.ViewSet):
    permission_classes = (IsAuthenticated,)

    def create(self, request, pk):
        queryset = Regular.objects.all()
        regular = get_object_or_404(queryset, pk=pk)
        representative = request.user
        queryset = Application.objects.all()
        application = get_object_or_404(queryset, regular=regular, representative=representative, status=False)
        Application.objects.filter(regular=regular, representative=representative, status=False).update(status=True)
        logger.info(f'User (id = {request.user.id}) upload results for tournament (id = {regular.id})')
        return file_parser(request, regular)


class CupUploadViewSet(viewsets.ViewSet):
    permission_classes = (IsAuthenticated,)

    def create(self, request, pk):
        queryset = Cup.objects.filter(posted_by=request.user)
        cup = get_object_or_404(queryset, pk=pk)
        logger.info(f'User (id = {request.user.id}) upload results for tournament (id = {cup.id})')
        return file_parser(request, cup)


class TournamentResultsViewSet(viewsets.ViewSet):
    def list(self, request, pk):
        queryset = TournamentBaseModel.objects.all()
        tournament = get_object_or_404(queryset, pk=pk)
        queryset = TournamentCompetitorsTeams.objects.filter(tournament=tournament)
        serializer = TournamentCompetitorsTeamsSerializer(queryset, many=True)
        return Response(serializer.data)